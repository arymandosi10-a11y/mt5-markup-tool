import streamlit as st
import pandas as pd
import requests
import openpyxl
from io import BytesIO

# ============================================================
# PAGE CONFIG + WIDE LAYOUT
# ============================================================
st.set_page_config(page_title="MarkupX – Market Cost Engine", layout="wide")

st.markdown(
    """
    <style>
      .block-container {max-width: 1800px; padding-top: 1.1rem; padding-bottom: 2rem;}
      .stDataFrame {border-radius: 12px;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("📊 MarkupX – Real Market Markup & Cost Engine")
st.caption("Real market prices + Profit-currency → USD conversion | Forex • Indices • Metals • Energies")

# ============================================================
# Yahoo Finance quote endpoint (NO yfinance library needed)
# ============================================================
@st.cache_data(ttl=30)
def yahoo_quote(symbol: str):
    url = f"https://query1.finance.yahoo.com/v7/finance/quote?symbols={symbol}"
    r = requests.get(url, timeout=12)
    r.raise_for_status()
    data = r.json()
    res = data.get("quoteResponse", {}).get("result", [])
    return res[0] if res else None

@st.cache_data(ttl=30)
def get_market_price_from_yahoo(yahoo_symbol: str):
    try:
        q = yahoo_quote(yahoo_symbol)
        if not q:
            return None
        p = q.get("regularMarketPrice", None)
        return float(p) if p is not None else None
    except Exception:
        return None

@st.cache_data(ttl=60)
def fx_to_usd_rate(ccy: str):
    if not ccy:
        return 1.0
    ccy = str(ccy).upper().strip()
    if ccy == "USD":
        return 1.0
    # Yahoo FX format: EURUSD=X
    p = get_market_price_from_yahoo(f"{ccy}USD=X")
    return p

# ============================================================
# Broker symbol -> Yahoo mapping (edit anytime)
# ============================================================
DEFAULT_YAHOO_MAP = {
    # Metals
    "XAUUSD": "XAUUSD=X",
    "XAGUSD": "XAGUSD=X",

    # Oil / Energies
    "USOIL": "CL=F",   # WTI Crude futures
    "UKOIL": "BZ=F",   # Brent futures
    "NGAS": "NG=F",

    # Indices (common broker CFD names)
    "NAS100": "^NDX",
    "US100": "^NDX",
    "SP500": "^GSPC",
    "US500": "^GSPC",
    "DJIUSD": "^DJI",
    "US30": "^DJI",
    "WS30": "^DJI",

    # Add more if your broker uses different names
}

def clean_symbol(s: str) -> str:
    return str(s).strip()

def infer_fx_yahoo_symbol(symbol_name: str):
    """
    If symbol looks like FX pair (6 letters), Yahoo format is EURNZD=X
    """
    s = clean_symbol(symbol_name).upper()
    if len(s) == 6 and s.isalpha():
        return f"{s}=X"
    return None

def resolve_price_source(row, symbol_col_name: str):
    """
    Priority:
    1) If file has Price_Source column -> use it
    2) If FX pair -> infer EURNZD=X etc
    3) Use DEFAULT_YAHOO_MAP for common CFD names
    4) Fallback to the symbol itself (may work for some)
    """
    # optional column in sheet
    for k in ["Price_Source", "Price Source", "Yahoo", "Yahoo_Symbol", "Yahoo Symbol"]:
        if k in row.index and pd.notna(row[k]) and str(row[k]).strip() != "":
            return str(row[k]).strip()

    sym = clean_symbol(row[symbol_col_name]).upper()

    fx = infer_fx_yahoo_symbol(sym)
    if fx:
        return fx

    if sym in DEFAULT_YAHOO_MAP:
        return DEFAULT_YAHOO_MAP[sym]

    return sym

def point_size_from_digits(digits) -> float:
    try:
        d = int(digits)
        return 10 ** (-d)
    except Exception:
        return None

# ============================================================
# SIDEBAR SETTINGS
# ============================================================
st.sidebar.header("⚙️ Settings")

default_lots = st.sidebar.number_input("Lots (default)", min_value=0.01, value=1.00, step=0.01)
default_markup_points = st.sidebar.number_input("Markup (points)", min_value=0.0, value=20.0, step=1.0)

lp_rate_per_1m_per_side = st.sidebar.number_input(
    "LP rate ($ per 1M USD notional per side)", min_value=0.0, value=7.0, step=0.5
)
lp_sides = st.sidebar.selectbox("LP sides", options=[1, 2], index=1)

st.sidebar.divider()
st.sidebar.subheader("IB Commission (NO sides)")

ib_type = st.sidebar.selectbox("Type", ["None", "Fixed ($/lot)", "Point-wise"], index=0)
ib_fixed_per_lot = st.sidebar.number_input("Fixed ($ per lot)", min_value=0.0, value=10.0, step=1.0)
ib_pointwise_points = st.sidebar.number_input(
    "Point-wise points (PointValue_USD * points * lots)", min_value=0.0, value=20.0, step=1.0
)

st.sidebar.divider()
st.sidebar.subheader("Display")
max_rows = st.sidebar.number_input("Show first N rows", min_value=10, max_value=5000, value=500, step=50)

# ============================================================
# FILE UPLOAD
# ============================================================
uploaded_file = st.file_uploader("📤 Upload your Excel (Book2.xlsx / symbol sheet)", type=["xlsx"])

if uploaded_file is None:
    st.warning("Please upload your Excel file to continue.")
    st.stop()

# ============================================================
# LOAD EXCEL
# ============================================================
try:
    df = pd.read_excel(uploaded_file, sheet_name=0)
except Exception as e:
    st.error(f"Failed to read Excel. Error: {e}")
    st.stop()

df.columns = [str(c).strip() for c in df.columns]

# ============================================================
# ACCEPT YOUR Book2.xlsx STRUCTURE
# Book2 columns are:
#   Symbol Name, Current Price, Digits, Profit Currency, Contract Size
# We will compute PointSize and PointValue_Profit_perlot automatically.
# Also: Lots can be optional column; otherwise sidebar default_lots is used.
# ============================================================
# Column aliases
symbol_col = None
for c in ["Symbol Name", "Symbol", "SymbolName", "Instrument", "Instrument Name"]:
    if c in df.columns:
        symbol_col = c
        break

profit_ccy_col = None
for c in ["Profit Currency", "Profit_Currency", "ProfitCurrency", "Currency"]:
    if c in df.columns:
        profit_ccy_col = c
        break

digits_col = None
for c in ["Digits", "Digit"]:
    if c in df.columns:
        digits_col = c
        break

contract_col = None
for c in ["Contract Size", "ContractSize", "Contract"]:
    if c in df.columns:
        contract_col = c
        break

current_price_col = None
for c in ["Current Price", "Price", "CurrentPrice"]:
    if c in df.columns:
        current_price_col = c
        break

# Optional per-row lots / per-row markup overrides
lots_col = "Lots" if "Lots" in df.columns else None
markup_override_col = "Markup_Points" if "Markup_Points" in df.columns else None  # optional

missing = []
if symbol_col is None: missing.append("Symbol Name (or Symbol)")
if profit_ccy_col is None: missing.append("Profit Currency")
if digits_col is None: missing.append("Digits")
if contract_col is None: missing.append("Contract Size")

if missing:
    st.error("Your Excel is missing required columns: " + ", ".join(missing))
    st.stop()

# ============================================================
# Build live report
# Rules you requested:
# - Indices/Metals/Energies may not have base currency -> we DO NOT use base at all
# - Use ONLY Profit Currency -> USD conversion for all calculations
# - Real time price from market source (Yahoo). If Yahoo not available, fallback to Current Price in sheet.
# - Brokerage = Markup_USD - LP_Commission_USD - IB_Commission_USD
# ============================================================
out = []
warnings = []

for i, r in df.iterrows():
    sym = clean_symbol(r[symbol_col])
    sym_u = sym.upper()

    profit_ccy = str(r[profit_ccy_col]).upper().strip() if pd.notna(r[profit_ccy_col]) else "USD"
    digits = r[digits_col]
    contract_size = float(r[contract_col]) if pd.notna(r[contract_col]) else 0.0

    lots = float(r[lots_col]) if (lots_col and pd.notna(r.get(lots_col))) else float(default_lots)

    # Markup points per symbol (optional), else sidebar default
    if markup_override_col and pd.notna(r.get(markup_override_col)) and str(r.get(markup_override_col)).strip() != "":
        try:
            markup_points = float(r.get(markup_override_col))
        except Exception:
            markup_points = float(default_markup_points)
    else:
        markup_points = float(default_markup_points)

    # PointSize and PointValue (Profit currency)
    ps = point_size_from_digits(digits)
    if ps is None:
        ps = 0.0
        warnings.append(f"{sym}: Digits invalid. PointSize set to 0.")
    point_value_profit_per_lot = contract_size * ps  # matches your indices example (digits=1 => ps=0.1)

    # FX profit -> USD
    fx = fx_to_usd_rate(profit_ccy)
    if fx is None:
        fx = 1.0
        warnings.append(f"{sym}: FX rate missing for {profit_ccy}. Used 1.0 fallback.")

    point_value_usd_per_lot = point_value_profit_per_lot * fx

    # Real market price
    yahoo_src = resolve_price_source(r, symbol_col)
    live_price = get_market_price_from_yahoo(yahoo_src)

    # fallback to sheet current price if yahoo fails
    sheet_price = None
    if current_price_col and pd.notna(r.get(current_price_col)):
        try:
            sheet_price = float(r.get(current_price_col))
        except Exception:
            sheet_price = None

    price_used = live_price if live_price is not None else sheet_price

    if price_used is None:
        price_used = 0.0
        warnings.append(f"{sym}: Market price not found (Yahoo: {yahoo_src}) and no Current Price usable. Price=0.")

    # Notional in Profit currency (for FX/CFD this is commonly price * contract * lots)
    notional_profit = price_used * contract_size * lots
    notional_usd = notional_profit * fx

    # Markup
    markup_profit = markup_points
    markup_usd = point_value_usd_per_lot * markup_points * lots

    # LP commission (per million USD notional * sides)
    lp_commission_usd = (notional_usd / 1_000_000.0) * float(lp_rate_per_1m_per_side) * int(lp_sides)

    # IB commission (NO sides)
    if ib_type == "Fixed ($/lot)":
        ib_commission_usd = float(ib_fixed_per_lot) * lots
    elif ib_type == "Point-wise":
        ib_commission_usd = point_value_usd_per_lot * float(ib_pointwise_points) * lots
    else:
        ib_commission_usd = 0.0

    # Brokerage formula (your requirement)
    brokerage_usd = markup_usd - lp_commission_usd - ib_commission_usd

    out.append({
        "Symbol": sym,
        "Profit_Currency": profit_ccy,
        "Price_Source_Used": yahoo_src,
        "Price_Live": live_price if live_price is not None else "",
        "Price_Sheet": sheet_price if sheet_price is not None else "",
        "Price_Used": price_used,

        "Digits": int(digits) if pd.notna(digits) else "",
        "PointSize": ps,
        "Contract_Size": contract_size,
        "Lots": lots,

        "FX_Profit_to_USD": fx,

        # Yellow style columns / key outputs
        "PointValue_Profit_perlot": point_value_profit_per_lot,
        "PointValue_USD_perlot": point_value_usd_per_lot,

        "Markup_Points": markup_points,
        "Markup_USD": markup_usd,

        "Notional_Profit": notional_profit,
        "Notional_USD": notional_usd,

        "LP_Commission_USD": lp_commission_usd,
        "IB_Commission_USD": ib_commission_usd,

        "Brokerage_USD": brokerage_usd
    })

out_df = pd.DataFrame(out)

# ============================================================
# DISPLAY
# ============================================================
st.subheader("📈 Live Cost Calculation")

default_cols = [
    "Symbol", "Profit_Currency", "Price_Used",
    "PointValue_USD_perlot", "Markup_Points", "Markup_USD",
    "Notional_USD", "LP_Commission_USD", "IB_Commission_USD", "Brokerage_USD",
    "Price_Source_Used"
]

selected_cols = st.multiselect(
    "Select columns to display",
    options=list(out_df.columns),
    default=[c for c in default_cols if c in out_df.columns],
)

view_df = out_df[selected_cols].head(int(max_rows)) if selected_cols else out_df.head(int(max_rows))
st.dataframe(view_df, use_container_width=True, hide_index=True)

if warnings:
    with st.expander("⚠️ Warnings / Missing data"):
        for w in warnings[:250]:
            st.write("•", w)
        if len(warnings) > 250:
            st.write(f"...and {len(warnings)-250} more")

# ============================================================
# EXPORT EXCEL WITH FORMULAS (visible when you click cells)
# NOTE:
# - Live price & FX are values (not formulas) because Excel can't fetch live Yahoo without add-ins.
# - All calculations (yellow columns) are formulas referencing those cells.
# ============================================================
def export_excel_with_formulas(report_df: pd.DataFrame):
    wb = openpyxl.Workbook()
    ws_set = wb.active
    ws_set.title = "Settings"

    ws_set["A1"] = "Default_Lots"; ws_set["B1"] = float(default_lots)
    ws_set["A2"] = "Default_Markup_Points"; ws_set["B2"] = float(default_markup_points)
    ws_set["A3"] = "LP_Rate_per_1M_per_side"; ws_set["B3"] = float(lp_rate_per_1m_per_side)
    ws_set["A4"] = "LP_Sides"; ws_set["B4"] = int(lp_sides)
    ws_set["A5"] = "IB_Type"; ws_set["B5"] = ib_type
    ws_set["A6"] = "IB_Fixed_per_lot"; ws_set["B6"] = float(ib_fixed_per_lot)
    ws_set["A7"] = "IB_Pointwise_Points"; ws_set["B7"] = float(ib_pointwise_points)

    ws = wb.create_sheet("Report")
    cols = list(report_df.columns)
    ws.append(cols)

    def col_letter(n):
        s = ""
        while n:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    cidx = {name: i + 1 for i, name in enumerate(cols)}

    # We'll set formulas for:
    # PointSize, PointValue_Profit_perlot, PointValue_USD_perlot,
    # Markup_USD, Notional_Profit, Notional_USD, LP_Commission_USD, IB_Commission_USD, Brokerage_USD
    for i, row in report_df.iterrows():
        excel_row = i + 2
        ws.append([row.get(c, "") for c in cols])

        def cell(col_name):
            return f"{col_letter(cidx[col_name])}{excel_row}"

        # references
        digits_cell = cell("Digits")
        contract_cell = cell("Contract_Size")
        lots_cell = cell("Lots")
        price_used_cell = cell("Price_Used")
        fx_cell = cell("FX_Profit_to_USD")
        markup_points_cell = cell("Markup_Points")
        pv_profit_cell = cell("PointValue_Profit_perlot")
        pv_usd_cell = cell("PointValue_USD_perlot")
        notional_profit_cell = cell("Notional_Profit")
        notional_usd_cell = cell("Notional_USD")
        markup_usd_cell = cell("Markup_USD")
        lp_cell = cell("LP_Commission_USD")
        ib_cell = cell("IB_Commission_USD")
        brokerage_cell = cell("Brokerage_USD")

        # PointSize = 10^-Digits
        ws[cell("PointSize")].value = f"=10^-{digits_cell}"

        # PointValue_Profit_perlot = Contract_Size * PointSize
        ws[pv_profit_cell].value = f"={contract_cell}*{cell('PointSize')}"

        # PointValue_USD_perlot = PointValue_Profit_perlot * FX
        ws[pv_usd_cell].value = f"={pv_profit_cell}*{fx_cell}"

        # Markup_USD = PointValue_USD_perlot * Markup_Points * Lots
        ws[markup_usd_cell].value = f"={pv_usd_cell}*{markup_points_cell}*{lots_cell}"

        # Notional_Profit = Price_Used * Contract_Size * Lots
        ws[notional_profit_cell].value = f"={price_used_cell}*{contract_cell}*{lots_cell}"

        # Notional_USD = Notional_Profit * FX
        ws[notional_usd_cell].value = f"={notional_profit_cell}*{fx_cell}"

        # LP_Commission_USD = (Notional_USD/1,000,000) * LP_Rate * LP_Sides
        ws[lp_cell].value = f"=({notional_usd_cell}/1000000)*Settings!$B$3*Settings!$B$4"

        # IB_Commission_USD:
        # Fixed ($/lot) => Lots * Settings!B6
        # Point-wise => PointValue_USD_perlot * Settings!B7 * Lots
        ws[ib_cell].value = (
            f"=IF(Settings!$B$5=\"Fixed ($/lot)\",{lots_cell}*Settings!$B$6,"
            f"IF(Settings!$B$5=\"Point-wise\",{pv_usd_cell}*Settings!$B$7*{lots_cell},0))"
        )

        # Brokerage_USD = Markup_USD - LP - IB
        ws[brokerage_cell].value = f"={markup_usd_cell}-{lp_cell}-{ib_cell}"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

st.subheader("⬇️ Download Excel (with formulas)")

st.download_button(
    "Download MarkupX Report (Formulas Included)",
    data=export_excel_with_formulas(out_df),
    file_name="MarkupX_Report_With_Formulas.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
